import os #---> To import the os to load all files
from openpyxl import load_workbook
folder = "." #---> current folder
#So the thing is that I couldn't have .xlsx type of files in my main folder , so pardon me for that. 
#I don't know if I made my point come across. 
#To get all the files in the folder.
all_files = os.listdir(folder)#A list of all files
xlsx_files = [] #Again a list
for file in all_files: #---> To get all files in all_files, which now contains all the files of our type in the folder
    if file.endswith(".xlsx"):
        xlsx_files.append(file) #-->Appended to the list 
print("Found", len(xlsx_files), "Excel files")
print("=" * 40)
for xlsx in xlsx_files:
    filepath = os.path.join(folder, xlsx)
    wb = load_workbook(filepath)
    props = wb.properties
    if props.creator:
        creator = props.creator
    else:
        creator = "Unknown"
    if props.title:
        title = props.title
    else:
        title = "Unknown"
    if props.created:
        created = props.created
    else:
        created = "Unknown"
    sheets = len(wb.sheetnames)
    print("File:", xlsx)
    print("  Creator:", creator)
    print("  Title:", title)
    print("  Created:", created)
    print("  Sheets:", sheets)
    print("-" * 30)
